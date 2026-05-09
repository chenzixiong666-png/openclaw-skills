#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import sys

# 打开 Excel 文件
file_path = r"C:\Users\woan\.openclaw\media\inbound\测试长运指标用例---bf3c320e-3746-4dba-97dc-550086c3149c.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print("="*120)
    print(f"📊 工作表: {ws.title}")
    print(f"📈 行数: {ws.max_row}, 列数: {ws.max_column}")
    print("="*120 + "\n")
    
    # 读取表头
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        headers.append(str(cell_value) if cell_value else "")
    
    print("📋 表头：")
    for idx, header in enumerate(headers, 1):
        print(f"  列 {idx}: {header}")
    
    print("\n" + "="*120)
    print(f"✅ 数据行数: {ws.max_row - 1}\n")
    print("="*120 + "\n")
    
    # 读取所有数据
    all_data = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx in range(1, len(headers) + 1):
            header = headers[col_idx - 1]
            cell_value = ws.cell(row_idx, col_idx).value
            row_data[header] = str(cell_value) if cell_value else ""
        all_data.append(row_data)
        
    # 打印前15行
    for idx, row in enumerate(all_data[:15], 1):
        print(f"用例 {idx}:")
        for key, value in row.items():
            if value:
                print(f"  {key}: {value}")
        print()
    
    print("="*120)
    print(f"✨ 总计 {len(all_data)} 个测试用例")
    
except Exception as e:
    print(f"错误: {e}")
    import traceback
    traceback.print_exc()
